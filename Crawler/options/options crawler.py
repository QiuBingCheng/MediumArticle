# -*- coding: utf-8 -*-
"""
Created on Sun Aug  2 07:06:38 2020

@author: Jerry
"""
import warnings
import requests
from openpyxl import load_workbook
from bs4 import BeautifulSoup
from datetime import datetime,timedelta
import time
warnings.filterwarnings("ignore")

#get excel last_date
def get_last_date(ws):
    print("get excel last date...")
    for i in range(2,ws.max_row+2):
        #inspect date
        if not ws[f"A{i}"].value:
            #remove weekday
            last_date = ws["A"+str(i-1)].value.split(" ")[0]
            #new i and last date
            print(last_date)
            return (i-1,last_date)
  
def get_daily_quotes(date,market_code):
    """
    market_code
    1=盤後
    0=一般
    """
    url = "https://www.taifex.com.tw/cht/3/futDailyMarketReport"
    myobj = {'queryDate': date, "MarketCode":market_code,"commodity_id":"TX","queryType":2}
    response = requests.post(url, data = myobj)
    soup = BeautifulSoup(response.text,features="html.parser")
    table = soup.find("table",class_="table_f")
    if not table:return None
    row = table.find_all("tr")[1].find_all("td")
    row = [r.text.strip() for r in row]
    return row

def get_date_during_the_period(start,end):
    duration = (end-start).days+1  #相差幾天(包含起始日)
    if duration>0:
        return [start+timedelta(days=d) for d in range(duration)]
    else:
        return None
    
def get_daily_quotes_by_date_list(dates):
    weekday_mapping = {0:"(週一)",1:"(週二)",2:"(週三)",3:"(週四)",4:"(週五)",5:"(週六)",6:"(週日)"}
    
    daily_quotes = []
    for date in dates:
        d = {}
        weekday = weekday_mapping[date.weekday()]
        date_str = date.strftime("%Y/%m/%d")
        full_date = date_str+" "+ weekday
        
        d["date"] = full_date
        d["after_hours"] = get_daily_quotes(date_str,1)
        d["normal"] = get_daily_quotes(date_str,0)
        
        time.sleep(0.5)
        
        if (not d["after_hours"]) | (not d["normal"]):
            continue
        daily_quotes.append(d)
    return daily_quotes

def record_to_excel(ws,row,daily_quotes):
    #record to excel
    for index,quotes in enumerate(daily_quotes):
        #欲填入的列數
        record_row = row+index
        
        ws[f"A{record_row}"] = quotes["date"]
        print(ws[f"A{record_row}"].value)
    
        #盤後
        ws[f"C{record_row}"] = int(quotes["after_hours"][2])
        ws[f"D{record_row}"] = int(quotes["after_hours"][3])
        ws[f"E{record_row}"] = int(quotes["after_hours"][4])
        ws[f"F{record_row}"] = int(quotes["after_hours"][5])
        
        #一般
        ws[f"G{record_row}"] = int(quotes["normal"][2])
        ws[f"H{record_row}"] = int(quotes["normal"][3])
        ws[f"I{record_row}"] = int(quotes["normal"][4])
        ws[f"J{record_row}"] = int(quotes["normal"][5])
        
        quote_change = "".join([s for s in quotes["normal"][6] if (s.isdigit() or s=="-")])
        ws[f"B{record_row}"] = int(quote_change)
        ws[f"K{record_row}"] = int(quotes["normal"][11])
        
#日期
wb = load_workbook("options.xlsx")
ws = wb[wb.sheetnames[0]]
last_date_row,last_date = get_last_date(ws)
last_date = datetime.strptime(last_date, "%Y/%m/%d")

today = datetime.today()
date_list = get_date_during_the_period(last_date+timedelta(days=1),today)
if date_list:
    daily_quotes = get_daily_quotes_by_date_list(date_list)
    record_to_excel(ws,last_date_row+1,daily_quotes)
    wb.save("options.xlsx")
else:
    print("Already up to date")
    
print("OK")

